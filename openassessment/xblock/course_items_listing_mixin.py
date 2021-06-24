"""
The mixin with handlers for the course ora blocks listing view.

"""

import json

from webob import Response
from xblock.core import XBlock
from openassessment.xblock.staff_area_mixin import require_course_staff
from django.utils.translation import ugettext as _
from opaque_keys.edx.keys import CourseKey
try:
    from openpyxl import Workbook
    from openpyxl.writer.excel import save_virtual_workbook
    from openpyxl.styles import PatternFill, Font
except ImportError:
    pass


class CourseItemsListingMixin:
    """
    The mixin with handlers for the course ora blocks listing view.

    """

    @XBlock.handler
    @require_course_staff('STAFF_AREA')
    def get_ora2_responses(self, request, suffix=''):  # pylint: disable=unused-argument
        """
        Get information about all ora2 blocks in the course with response count for each step.

        """
        # Import is placed here to avoid model import at project startup.
        from openassessment.data import OraAggregateData
        responses = OraAggregateData.collect_ora2_responses(str(self.course_id))
        return Response(json.dumps(responses), content_type='application/json', charset='UTF-8')

    @XBlock.handler
    @require_course_staff('STAFF_AREA')
    def download_ora2_responses(self, request, suffix=''):  # pylint: disable=unused-argument
        from openassessment.data import OraAggregateData
        ora_responses = OraAggregateData.collect_ora2_responses(str(self.course_id))
        store = self.runtime.modulestore
        parents = {}
        course_key = CourseKey.from_string(self.course_id)
        report_name = self.course_id[10:].replace('+', ' ') + ' - open responses'

        sheet_name = 'Open Responses'
        workbook = Workbook()
        workbook.active.title = sheet_name
        sheet = workbook[sheet_name]
        sheet.append([
            'Section',
            'Subsection',
            'Unit Name',
            'Assessment',
            'Total Responses',
            'Training',
            'Peer',
            'Self',
            'Waiting',
            'Staff',
            'Final Grade Received'
        ])

        final_num = 0
        final_total_val = 0
        final_training_val = 0
        final_peer_val = 0
        final_self_val = 0
        final_waiting_val = 0
        final_staff_val = 0
        final_done_val = 0

        with store.bulk_operations(course_key):
            openassessment_blocks = store.get_items(
                course_key, qualifiers={'category': 'openassessment'}
            )
            # filter out orphaned openassessment blocks
            openassessment_blocks = [
                block for block in openassessment_blocks if block.parent is not None
            ]
            for block in openassessment_blocks:
                result_item_id = str(block.location.replace(version_guid=None, branch=None))
                has_staff_assessment = 'staff-assessment' in block.assessment_steps
                if result_item_id not in parents:
                    vert, seq, chapter = _get_full_path_names(block, store)
                    if vert and seq and chapter:
                        parents[result_item_id] = (vert, seq, chapter)
                    else:
                        continue

                assessment_name = _("Team") + " : " + block.display_name if block.teams_enabled else block.display_name
                chapter_name = parents[result_item_id][2].display_name
                seq_name = parents[result_item_id][1].display_name
                parent_name = parents[result_item_id][0].display_name
                ora_item_dict = ora_responses.get(result_item_id, {})

                training_val = ora_item_dict.get('training', 0)
                peer_val = ora_item_dict.get('peer', 0)
                self_val = ora_item_dict.get('self', 0)
                waiting_val = ora_item_dict.get('waiting', 0)
                staff_val = ora_item_dict.get('staff', 0)
                done_val = ora_item_dict.get('done', 0)

                if has_staff_assessment:
                    staff_val = waiting_val
                    waiting_val = 0
                total_val = training_val + peer_val + self_val + waiting_val + staff_val + done_val

                final_num = final_num + 1
                final_total_val = final_total_val + total_val
                final_training_val = final_training_val + training_val
                final_peer_val = final_peer_val + peer_val
                final_self_val = final_self_val + self_val
                final_waiting_val = final_waiting_val + waiting_val
                final_staff_val = final_staff_val + staff_val
                final_done_val = final_done_val + done_val

                sheet.append([
                    chapter_name,
                    seq_name,
                    parent_name,
                    assessment_name,
                    total_val,
                    training_val,
                    peer_val,
                    self_val,
                    waiting_val,
                    staff_val,
                    done_val
                ])

        sheet.append([
            final_num,
            final_num,
            final_num,
            final_num,
            final_total_val,
            final_training_val,
            final_peer_val,
            final_self_val,
            final_waiting_val,
            final_staff_val,
            final_done_val
        ])
        for cell in sheet[sheet.max_row]:
            cell.fill = PatternFill('solid', fgColor='B8CCE4')
            cell.font = Font(b=True)
        content = save_virtual_workbook(workbook)
        return Response(content, content_disposition='attachment; filename=' +  report_name + '.xlsx')


def _get_full_path_names(block, store):
    vert, seq, chapter = None, None, None
    if block.parent:
        vert = store.get_item(block.parent)
    if vert.parent:
        seq = store.get_item(vert.parent)
    if seq.parent:
        chapter = store.get_item(seq.parent)
    return vert, seq, chapter
